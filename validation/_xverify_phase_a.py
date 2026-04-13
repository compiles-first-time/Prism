"""
Phase A: Cross-reference sweep between the workbook and the expanded specs.

Produces three reports:
1. Broken references — SR IDs referenced by BRs but not defined in any expanded spec
2. Unreferenced SRs — SR IDs defined in specs but not referenced by any BR
3. BR-to-SR mapping table — for manual review of semantic alignment

Writes results to _xverify_phase_a_report.txt
"""
import openpyxl
import re
import pathlib
from collections import defaultdict

WORKBOOK_PATH = r"D:\Projects\IDEA\working\Platform_Requirements_and_Exceptions.xlsx"
SPECS_DIR = pathlib.Path(r"D:\Projects\IDEA\explore\002-spec-expansion\expanded-specs")
# Kernel V6 supplement files live alongside the original specs in build/spec/
KERNEL_SPECS_DIR = pathlib.Path(r"D:\Projects\IDEA\build\spec")
REPORT_PATH = r"D:\Projects\IDEA\working\_xverify_phase_a_report.txt"

# SR ID regex — matches SR_GOV_47, SR_LLM_50, SR_CONN_11_BE-01, SR_GOV_46_REVIEW, SR_V2_01, SR_SCALE_25_BE-01
# First segment after SR_ is alphanumeric (allows V2) — must have at least one letter
SR_PATTERN = re.compile(r'SR_[A-Z][A-Z0-9]*_\d+(?:_[A-Za-z0-9-]+)?')

# ---------------------------------------------------------------------------
# Step 1: Extract every SR reference from the workbook
# ---------------------------------------------------------------------------
print("Loading workbook...")
wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

# workbook_map: SR_id -> list of (sheet_name, br_id)
workbook_map = defaultdict(list)
# br_map: br_id -> (sheet_name, list of SR refs)
br_map = {}

total_brs = 0
total_refs = 0

for sheet_name in wb.sheetnames:
    if sheet_name == "Index":
        continue
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        br_id = row[0]
        total_brs += 1
        sr_refs_cell = row[3] if len(row) > 3 else None
        if not sr_refs_cell:
            br_map[br_id] = (sheet_name, [])
            continue
        # Filter out NEW markers and notes
        if "NEW" in str(sr_refs_cell) and "SR_" not in str(sr_refs_cell).split("NEW")[0]:
            # Entries like "NEW — no SR yet (gap from Nick's workbook)"
            br_map[br_id] = (sheet_name, ["__NEW__"])
            continue
        refs = SR_PATTERN.findall(str(sr_refs_cell))
        br_map[br_id] = (sheet_name, refs)
        for r in refs:
            workbook_map[r].append((sheet_name, br_id))
            total_refs += 1

print(f"  BRs scanned: {total_brs}")
print(f"  SR references found: {total_refs}")
print(f"  Unique SR IDs referenced: {len(workbook_map)}")

# ---------------------------------------------------------------------------
# Step 2: Extract every defined SR ID from the expanded specs
# ---------------------------------------------------------------------------
print("\nScanning expanded specs...")

defined_srs = set()
# sr_definitions: SR_id -> (spec_file, line_num, row_text)
sr_definitions = {}

# SR row pattern — left-most cell in a markdown table row, backticked
SR_ROW_PATTERN = re.compile(r'^\| `(SR_[A-Z][A-Z0-9]*_\d+(?:_[A-Za-z0-9-]+)?)`', re.MULTILINE)

spec_dirs = [
    (SPECS_DIR, "*.md"),           # Original 14 expanded specs
    (KERNEL_SPECS_DIR, "*-kernel-v6.md"),  # Kernel V6 supplement files
]

for spec_dir, pattern in spec_dirs:
    if not spec_dir.exists():
        print(f"  WARNING: {spec_dir} does not exist — skipping")
        continue
    for spec_file in sorted(spec_dir.glob(pattern)):
        with open(spec_file, encoding="utf-8") as f:
            content = f.read()
        for match in SR_ROW_PATTERN.finditer(content):
            sr_id = match.group(1)
            defined_srs.add(sr_id)
            if sr_id not in sr_definitions:
                # Find the line
                line_num = content[:match.start()].count("\n") + 1
                # Get the row text (truncated)
                line_end = content.find("\n", match.start())
                row_text = content[match.start():line_end][:200] if line_end > 0 else ""
                sr_definitions[sr_id] = (spec_file.name, line_num, row_text)

print(f"  SRs defined across all specs (original + kernel): {len(defined_srs)}")

# ---------------------------------------------------------------------------
# Step 3: Compute broken references and unreferenced SRs
# ---------------------------------------------------------------------------
referenced_set = set(workbook_map.keys())
broken_refs = referenced_set - defined_srs
unreferenced_srs = defined_srs - referenced_set

# Also count coverage ratio
covered_count = len(referenced_set & defined_srs)

# ---------------------------------------------------------------------------
# Step 4: Write report
# ---------------------------------------------------------------------------
print(f"\nBroken references: {len(broken_refs)}")
print(f"Unreferenced SRs: {len(unreferenced_srs)}")
print(f"Covered (referenced + defined): {covered_count}")

with open(REPORT_PATH, "w", encoding="utf-8") as f:
    f.write("=" * 78 + "\n")
    f.write("PHASE A CROSS-VERIFICATION REPORT\n")
    f.write("=" * 78 + "\n\n")

    f.write("HIGH-LEVEL COUNTS\n")
    f.write("-" * 78 + "\n")
    f.write(f"  BRs in workbook:              {total_brs}\n")
    f.write(f"  SR references in workbook:    {total_refs}\n")
    f.write(f"  Unique SR IDs referenced:     {len(workbook_map)}\n")
    f.write(f"  SRs defined in specs:         {len(defined_srs)}\n")
    f.write(f"  Covered (ref + def):          {covered_count}\n")
    f.write(f"  Broken refs (ref not def):    {len(broken_refs)}\n")
    f.write(f"  Unreferenced SRs (def not ref): {len(unreferenced_srs)}\n\n")

    # -----------------------------------------------------------------------
    f.write("\n" + "=" * 78 + "\n")
    f.write("SECTION 1: BROKEN REFERENCES (workbook references that do not resolve)\n")
    f.write("=" * 78 + "\n\n")
    if not broken_refs:
        f.write("  NONE — all SR references resolve cleanly.\n")
    else:
        for sr_id in sorted(broken_refs):
            f.write(f"\n  {sr_id}\n")
            usages = workbook_map[sr_id]
            for sheet, br_id in usages:
                f.write(f"    Used by: [{sheet}] {br_id}\n")

    # -----------------------------------------------------------------------
    f.write("\n\n" + "=" * 78 + "\n")
    f.write("SECTION 2: UNREFERENCED SRs (spec SRs with no BR coverage)\n")
    f.write("=" * 78 + "\n")
    f.write("NOTE: This is not necessarily a problem. SRs are more granular than BRs.\n")
    f.write("      Many SRs are implementation details subsumed by a higher-level BR.\n")
    f.write("      This list is for manual review to decide which need BR coverage.\n\n")

    # Group unreferenced SRs by prefix for easier review
    by_prefix = defaultdict(list)
    for sr in sorted(unreferenced_srs):
        prefix = sr.split("_")[1]  # GOV, DM, CONN, INT, etc.
        by_prefix[prefix].append(sr)

    prefix_to_sheet = {
        "GOV": "1. Governance",
        "DS": "2. Forecasting",
        "CONN": "3. Connection",
        "INT": "4. Intelligence",
        "LLM": "5. LLM Routing",
        "DM": "6. Data Model",
        "UI": "7. Interface",
        "CAT": "8. Component Catalog",
        "SA": "9. Service Account Catalog",
        "FW": "10. Value Flywheel",
        "UU": "11. Unknown Unknowns",
        "V2": "12. V2 Handoff Contract",
        "SCALE": "13. Scalability Infrastructure",
    }

    for prefix in sorted(by_prefix.keys()):
        sheet = prefix_to_sheet.get(prefix, "?")
        srs = by_prefix[prefix]
        f.write(f"\n  [{sheet}] SR_{prefix}_* : {len(srs)} unreferenced\n")
        for sr in srs:
            spec_file, line_num, _ = sr_definitions.get(sr, ("?", 0, ""))
            f.write(f"    {sr}  ({spec_file}:{line_num})\n")

    # -----------------------------------------------------------------------
    f.write("\n\n" + "=" * 78 + "\n")
    f.write("SECTION 3: BR -> SR MAPPING (for semantic review)\n")
    f.write("=" * 78 + "\n\n")

    current_sheet = None
    for br_id in sorted(br_map.keys(), key=lambda x: (br_map[x][0], x)):
        sheet, refs = br_map[br_id]
        if sheet != current_sheet:
            current_sheet = sheet
            f.write(f"\n[{sheet}]\n")
        if not refs or refs == ["__NEW__"]:
            f.write(f"  {br_id}\n")
            f.write(f"    (no SR references — marked NEW or empty)\n")
        else:
            # Mark broken refs with !
            ref_str = ", ".join(
                f"!{r}" if r in broken_refs else r for r in refs
            )
            f.write(f"  {br_id}\n")
            f.write(f"    -> {ref_str}\n")

print(f"\nReport written to: {REPORT_PATH}")
